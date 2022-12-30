# coding=utf-8
import os
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, colors, Alignment, Font
path = r'E:\2022年项目\0707沙朗沟断面\slg\表格'.decode('utf-8')
os.chdir(path)
lst = []
for i,v,m in os.walk(path):
    for xls in m:
        lst.append(os.path.join(i, xls))

for csv in lst:
    print csv
    df = pd.read_csv(csv)
    dmh = re.findall(r'\w{2,10}-\d{3}',csv)
    title = '横断面成果表' + '(' + dmh[0].encode('utf-8') + ')'
    i = 1
    cnList = []
    for index,value in df.iterrows():
        a =  i,df.loc[index,'dis'],df.loc[index,'h'],df.loc[index,'y'],df.loc[index,'x']
        i+=1
        cnList.append(a)
    df2 = pd.DataFrame(cnList,columns=[title,'','','',''])
    df2.to_excel(path + '\\' + 'xlsx' + '\\' + '{}.xlsx'.format(dmh[0]),index=False)
    xlsx = path + '\\' + 'xlsx' + '\\' + '{}.xlsx'.format(dmh[0])
    wb = load_workbook(xlsx)
    ws = wb.worksheets[0]
    ws.merge_cells('A1:E1')
    ws.insert_rows(2,1)
    ws.cell(2,2,'{}'.format(dmh[0]))
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    border_set = Border(left=Side(style='thin', color=colors.BLACK),
                        right=Side(style='thin', color=colors.BLACK),
                        top=Side(style='thin', color=colors.BLACK),
                        bottom=Side(style='thin', color=colors.BLACK))
    for row in ws.rows:
        for cell in row:
            cell.border = border_set
            cell.alignment = Alignment(horizontal='center', vertical='center')
    for cell in ws['A']:
        cell.number_format = '0'
    for cell in ws['B']:
        cell.number_format = '0.00'
    for cell in ws['C']:
        cell.number_format = '0.00'
    for cell in ws['D']:
        cell.number_format = '0.0000'
    for cell in ws['E']:
        cell.number_format = '0.0000'
    QD = ws.cell(3,2).value
    print QD
    for cell in ws['B']:
        if type(cell.value) is float:
            cell.value  = cell.value -QD

    wb.save(path + '\\' + 'xlsx' + '\\' + '{}.xlsx'.format(dmh[0]))
