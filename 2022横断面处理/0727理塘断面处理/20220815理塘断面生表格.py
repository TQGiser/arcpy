#coding=utf-8
import arcpy
import pandas as pd
import math
from openpyxl import load_workbook
from openpyxl.styles import Border,Side,colors,Alignment,Font
path = r'E:\2022年项目\0815理塘断面表格'
okpath = path + '\\' + 'ok'
arcpy.env.workspace = path
dmx_all = r'E:\2022年项目\0815理塘断面表格\改理塘段面备注8.16\99.shp'
riverList = []

with arcpy.da.SearchCursor(dmx_all, 'RIVER') as yb:
    for row in yb:
        if row[0] not in riverList:
            riverList.append(row[0])

for river in riverList:
    arcpy.CreateFolder_management(okpath,river.encode('utf-8'))
    expression = "RIVER = '{}'".format(river.encode('utf-8'))
    txt = okpath + '\\' + river.encode('utf-8')+ '\\' + '河道断面采集散点数据.txt'
    cnList3 = []
    dmhList = []
    csv = okpath + '\\' + river.encode('utf-8')+ '\\' + '{}断面数据.csv'.format(river.encode('utf-8'))
    with open(txt.decode('utf-8'),'w') as f:
        j=0
        with arcpy.da.SearchCursor(dmx_all,['RIVER','dmh'],where_clause= expression) as yb:
            for row in yb:
                dmh = row[1]
                if row[1] not in dmhList:
                    dmhList.append(dmh)
        dmhList.sort()
        for dmh in dmhList:
                    expression2 = "dmh = '{}'".format(dmh)
                    xlsName =  '{}.xlsx'.format(dmh)
                    print '{}.xls'.format(dmh)
                    cnList2 = []
                    cn4 = [dmh,'','']
                    cnList3.append(cn4)
                    with arcpy.da.SearchCursor(dmx_all,['SHAPE@','RIVER','dmh','bz'],where_clause= expression2) as yb:
                        cnList1 = []
                        i=1
                        for row in yb:
                            cn1 = [i,row[0].centroid.X, row[0].centroid.Y]
                            cnList1.append(cn1)
                            ldj = round(math.sqrt(math.pow((row[0].centroid.X-cnList1[0][1]),2)+math.pow((row[0].centroid.Y-cnList1[0][2]),2)),2)
                            cn2 = [i,ldj,row[0].centroid.Z,row[3],row[0].centroid.X, row[0].centroid.Y]
                            cnList2.append(cn2)
                            cn3 = [ldj, row[0].centroid.Z,row[3]]
                            cnList3.append(cn3)
                            f.write(str(j))
                            f.write(",")
                            f.write(",")
                            f.write(str(row[0].centroid.X))
                            f.write(",")
                            f.write(str(row[0].centroid.Y))
                            f.write(",")
                            f.write(str(row[0].centroid.Z) + '\n')
                            # print i,ldj,row[0].centroid.Z,row[3],row[0].centroid.X, row[0].centroid.Y
                            i+=1
                            j+=1
                    df = pd.DataFrame(cnList2, columns=['点号', '另点距', '高程', '备注', 'X','Y'])
                    xlsx = okpath + '\\' + river.encode('utf-8') + '\\' + xlsName.encode('utf-8')
                    print xlsx,type(xlsx)
                    df.to_excel(xlsx.decode('utf-8'), index=False)

                    wb = load_workbook(xlsx.decode('utf-8'))
                    ws = wb.worksheets[0]
                    ws.insert_rows(2, 1)
                    ws['B2'] = dmh
                    ws.column_dimensions['C'].width = 15
                    ws.column_dimensions['E'].width = 15
                    ws.column_dimensions['F'].width = 15
                    border_set = Border(left=Side(style='thin', color=colors.BLACK),
                                        right=Side(style='thin', color=colors.BLACK),
                                        top=Side(style='thin', color=colors.BLACK),
                                        bottom=Side(style='thin', color=colors.BLACK))
                    for row in ws.rows:
                        for cell in row:
                            cell.border = border_set
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                    for cell in ws['E']:
                        cell.number_format = '0.0000'
                    for cell in ws['F']:
                        cell.number_format = '0.0000'
                    wb.save(xlsx.decode('utf-8'))
    df3 = pd.DataFrame(cnList3)
    df3.to_csv(csv.decode('utf-8'),index=False,encoding='GBK')
