# coding=utf-8
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, colors, Alignment, Font
pd.set_option('display.width',100)
path = r'C:\Users\Administrator\Desktop\新建文件夹'.decode('utf-8')
lst = []
os.chdir(path)
dfall = pd.DataFrame(columns= ['A','B','C','D','E','F','G','H','I'])
for i,v,m in os.walk(path):
    for xls in m:
        lst.append(os.path.join(i, xls))
for xls in lst:
    print xls
    df = pd.read_excel(xls, header=0)
    header = df.columns.tolist()[0]
    df.columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G','H','I']
    df.loc[0, 'A'] = header
    df.loc[1, 'A'] = '桩名（编号）'.decode('utf-8')
    df.loc[1, 'B'] = '里程'.decode('utf-8')
    df.loc[1, 'C'] = '所在位置（地名）'.decode('utf-8')
    df.loc[1, 'H'] = '高程（1985国家高程基准）'.decode('utf-8')
    dfall = dfall.append(df, ignore_index=True)
dfall.to_excel(r'C:\Users\Administrator\Desktop\新建文件夹\all.xlsx'.decode('utf-8'), index=False)
wb = load_workbook(r'C:\Users\Administrator\Desktop\新建文件夹\all.xlsx'.decode('utf-8'))
ws = wb.worksheets[0]

for cell in ws['A']:
    if cell.value is not None:
        if '管理线桩成果表'.decode('utf-8') in cell.value:
            a = cell.row
            ws.merge_cells('A{}:I{}'.format(a, a))
        if '告示牌成果表'.decode('utf-8') in cell.value:
            b = cell.row
            ws.merge_cells('A{}:I{}'.format(b, b))
            ws.delete_rows(b + 3)
            ws.delete_rows(b + 2)
            ws.delete_rows(b + 1)
            ws.delete_rows(b)
wb.save(r'C:\Users\Administrator\Desktop\新建文件夹\all.xlsx'.decode('utf-8'))