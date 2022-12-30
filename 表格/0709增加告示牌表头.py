# coding=utf-8
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, colors, Alignment, Font

path = r'D:\2021年项目\0708湖泊改库\xls加告示牌标题'.decode('utf-8')
xlslist = []
os.chdir(path)
for i, v, m in os.walk(path):
    for file in m:
        xlslist.append(os.path.join(i, file))
for xls in xlslist:
    s = '告示牌001'
    wb = load_workbook(xls)
    ws = wb.worksheets[0]
    name = ws['A1'].value
    for cell in ws['A']:
        if '告示牌001'.decode('utf-8') == cell.value:
            print xls
            a = cell.row
            ws.insert_rows(a, 3)
            wb.save(xls)
            ws['A{}'.format(a)].value = name.replace('管理线桩'.decode('utf-8'), '告示牌'.decode('utf-8'))
            ws['A{}'.format(a + 1)].value = ws['A2'].value
            ws['B{}'.format(a + 1)].value = ws['B2'].value
            ws['C{}'.format(a + 1)].value = ws['C2'].value
            ws['D{}'.format(a + 1)].value = ws['D2'].value
            ws['E{}'.format(a + 1)].value = ws['E2'].value
            ws['F{}'.format(a + 1)].value = ws['F2'].value
            ws['G{}'.format(a + 1)].value = ws['G2'].value
            ws['H{}'.format(a + 1)].value = ws['H2'].value
            ws['D{}'.format(a + 2)].value = ws['D3'].value
            ws['E{}'.format(a + 2)].value = ws['E3'].value
            ws['F{}'.format(a + 2)].value = ws['F3'].value
            ws['G{}'.format(a + 2)].value = ws['G3'].value
            ws.merge_cells('A{}:H{}'.format(a, a))
            ws.merge_cells('A{}:A{}'.format(a + 1, a + 2))
            ws.merge_cells('B{}:B{}'.format(a + 1, a + 2))
            ws.merge_cells('C{}:C{}'.format(a + 1, a + 2))
            ws.merge_cells('D{}:G{}'.format(a + 1, a + 1))
            ws.merge_cells('H{}:H{}'.format(a + 1, a + 2))



            ws.row_dimensions[a].height = 26
            border_set = Border(left=Side(style='thin', color=colors.BLACK),
                                right=Side(style='thin', color=colors.BLACK),
                                top=Side(style='thin', color=colors.BLACK),
                                bottom=Side(style='thin', color=colors.BLACK))
            for row in ws.rows:
                fontstyle2 = Font(size=11, bold=False)
                for cell in row:
                    cell.border = border_set
                    cell.number_format = '0.0000'
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = fontstyle2
            for cell in ws['H']:
                cell.number_format = '0.00'
            fontstyle = Font(size=20, bold=True)
            ws['A1'].font = fontstyle
            ws['A{}'.format(a)].font = fontstyle
            wb.save(xls)
