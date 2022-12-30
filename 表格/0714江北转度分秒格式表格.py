#coding=utf-8
# coding=utf-8
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, colors, Alignment, Font
path = r'D:\2021年项目\0713江北数据库\度分秒格式表格'.decode('utf-8')
xlslist = []
os.chdir(path)
for i, v, m in os.walk(path):
    for file in m:
        xlslist.append(os.path.join(i, file))
for xls in xlslist:
    wb = load_workbook(xls)
    ws = wb.worksheets[0]
    for cell in ws['D']:
        if cell.value is not None:
            if 'E' in cell.value:
                a = cell.value
                du = a.replace('E', '').replace('°'.decode('utf-8'), '').split('.')[0] + '°'.decode('utf-8')
                fen = str('%02d' % int(
                    float('0.' + a.replace('E', '').replace('°'.decode('utf-8'), '').split('.')[1]) * 60)) + "'".decode(
                    'utf-8')
                miao = str('%02d' % int(float('0.' + str(
                    float('0.' + a.replace('E', '').replace('°'.decode('utf-8'), '').split('.')[1]) * 60).split('.')[
                    1]) * 60))
                miao2 = '.'.decode('utf-8') + str('%03d' % round(float(str(float('0.' + str(
                    float('0.' + a.replace('E', '').replace('°'.decode('utf-8'), '').split('.')[1]) * 60).split('.')[
                    1]) * 60).split('.')[1]) / 1000000))[0:3] + '"'.decode('utf-8')
                dfm = du + fen + miao + miao2
                cell.value = dfm
    for cell in ws['E']:
        if cell.value is not None:
            if 'N' in cell.value:
                a = cell.value
                du = a.replace('N', '').replace('°'.decode('utf-8'), '').split('.')[0] + '°'.decode('utf-8')
                fen = str('%02d' % int(
                    float('0.' + a.replace('N', '').replace('°'.decode('utf-8'), '').split('.')[
                        1]) * 60)) + "'".decode(
                    'utf-8')
                miao = str('%02d' % int(float('0.' + str(
                    float('0.' + a.replace('N', '').replace('°'.decode('utf-8'), '').split('.')[
                        1]) * 60).split('.')[
                    1]) * 60))
                miao2 = '.'.decode('utf-8') + str('%03d' % round(float(str(float('0.' + str(
                    float('0.' + a.replace('N', '').replace('°'.decode('utf-8'), '').split('.')[
                        1]) * 60).split('.')[
                    1]) * 60).split('.')[1]) / 1000000))[0:3] + '"'.decode('utf-8')
                dfm = du + fen + miao + miao2
                cell.value = dfm
    wb.save(xls)