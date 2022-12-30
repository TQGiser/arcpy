# coding=utf-8
import arcpy
import os
import F
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border,Side,colors,Alignment,Font
path = r'E:\测试文件夹\A'.decode('utf-8')
nameList = []
df1 = pd.read_excel(r'E:\测试文件夹\A\OK\a.xlsx'.decode('utf-8'))
for name,value in df1.iterrows():
    nameList.append(df1.loc[name,'河流'.decode('utf-8')])
lst = []
os.chdir(path)
dfall = pd.DataFrame(columns= ['A','B','C','D','E','F','G','H','I'])
for name in nameList:
    for i,v,m in os.walk(path):
        for xls in m:
            if name in os.path.join(i, xls) and '划定成果表'.decode('utf-8') in os.path.join(i, xls):
                lst.append(os.path.join(i, xls))
for xls in lst:
    print xls
    riverName = xls.split('\划定'.decode('utf-8'))[0].split('A\\')[1]
    # print riverName
    df = pd.read_excel(xls, header=0)
    header = df.columns.tolist()[0]
    df.columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G','H','I']
    # df.loc[0, 'A'] = header
    # df.loc[1, 'A'] = '桩名（编号）'.decode('utf-8')
    # df.loc[1, 'B'] = '里程'.decode('utf-8')
    # df.loc[1, 'C'] = '所在位置（地名）'.decode('utf-8')
    # df.loc[1, 'H'] = '高程（1985国家高程基准）'.decode('utf-8')
    dfall = dfall.append(df, ignore_index=True)
dfall.to_excel(path + '\\' + 'all.xlsx', index=False)
wb = load_workbook(path + '\\' + 'all.xlsx')
ws = wb.worksheets[0]
for cell in ws['D']:
    if '经度'.decode('utf-8') ==cell.value:
        a = cell.row
        b = ((ws['A{}'.format(a+1)].value).split('-')[0]).encode('utf-8')
        ws['A{}'.format(a)].value = '桩名（编号）'.decode('utf-8')
        ws['B{}'.format(a)].value = '里程（km）'.decode('utf-8')
        ws['C{}'.format(a)].value = '所在位置（地名）'.decode('utf-8')

        ws['I{}'.format(a)].value = '备注'.decode('utf-8')
        ws.insert_rows(a,1)
        ws['A{}'.format(a)].value = ('四川省甘孜藏族自治州巴塘县{}管理线桩信息表'.format(b)).decode('utf-8')
        ws.merge_cells('A{}:I{}'.format(a,a))
# for cell in ws['A']:
#     if cell.value is not None:
#         if '管理线桩成果表'.decode('utf-8') in cell.value:
#             a = cell.row
#             ws.merge_cells('A{}:I{}'.format(a, a))
wb.save(path + '\\' + 'all.xlsx')