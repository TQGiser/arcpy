# coding=utf-8
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border,Side,colors,Alignment,Font
path = r'D:\2021年项目\0708湖泊改库\xls'.decode('utf-8')
mb = r'D:\2021年项目\0708湖泊改库\桩牌模板.xlsx'.decode('utf-8')
outpath = r'D:\2021年项目\0708湖泊改库\xlsOK'
temp_xls = r'D:\2021年项目\0708湖泊改库\temp_xls.xls'.decode('utf-8')
xlslist = []
os.chdir(path)
for i,v,m in os.walk(path):
    xlslist.append(m)
for xls in xlslist[0]:
    xlsPath = xls.encode('utf-8').split('.')[0]
    b = str(xlsPath)
    print b
    xName = b.decode('utf-8').split('县'.decode('utf-8'))[0] + '县'.decode('utf-8')
    lakeName = b.decode('utf-8').split('县'.decode('utf-8'))[-1]
    xDirs = outpath.decode('utf-8') + '\\' + xName
    lakeDirs = outpath.decode('utf-8') + '\\' + xName + '\\' + lakeName
    forder = os.path.exists(lakeDirs)
    if not forder:
        os.makedirs(lakeDirs)
    df = pd.read_excel(xls)
    df.drop('FID', axis=1, inplace=True)
    df2 = pd.read_excel(mb)
    df3 = pd.concat([df2,df],ignore_index=True)
    df3.to_excel(temp_xls,index=False)
    df4 = pd.read_excel(temp_xls, header=1)
    okxls = lakeDirs + '\\' + xls.split('.')[0] + '.xlsx'
    df4.to_excel(okxls,index=False)
    wb = load_workbook(okxls)
    ws = wb.worksheets[0]
    ws.insert_rows(0, 1)
    fontstyle = Font(size= 20,bold=True)
    ws.cell(1,1,'四川省甘孜州{}管理线桩信息表'.format(b))
    ws['A1'].font = fontstyle
    ws.merge_cells('A1:H1')
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 28
    ws.merge_cells('A2:A3')
    ws.merge_cells('B2:B3')
    ws.merge_cells('C2:C3')
    ws.merge_cells('D2:G2')
    ws.merge_cells('H2:H3')
    ws.row_dimensions[1].height = 26
    border_set = Border(left=Side(style='thin', color=colors.BLACK),
                        right=Side(style='thin', color=colors.BLACK),
                        top=Side(style='thin', color=colors.BLACK),
                        bottom=Side(style='thin', color=colors.BLACK))
    for row in ws.rows:
        for cell in row:
            cell.border = border_set
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal='center', vertical='center')
    for cell in ws['H']:
            cell.number_format = '0.00'
    wb.save(okxls)
