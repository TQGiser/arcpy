# coding=utf-8
import arcpy
import pandas as pd
import math
import os
import F
from openpyxl import load_workbook
from openpyxl.styles import Border,Side,colors,Alignment,Font
path = r'E:\测试文件夹\导SHP'
arcpy.env.workspace = path
mdbs = arcpy.ListFiles('*.mdb')
for mdb in mdbs:
    print mdb
    riverName = arcpy.Describe(mdb).name.replace('.mdb','')
    cnList = []
    dmap = path.decode('utf-8') + '\\' + mdb + '\\' + 'DLG' + '\\' + 'DMAP'
    with arcpy.da.SearchCursor(dmap,['NAME','NUM','COUNTY','TOWN','SHAPE@X','SHAPE@Y','ELEV','RuleID','SHAPE@','BANK']) as yb :
        for row in yb:
            name = row[0]
            if '右'.decode('utf-8') in name:
                pxzy = 'r'
            else:
                pxzy = 'l'
            lc = row[1]
            xzq = row[2] + row[3]
            e = 'E' + '%0.8f'%round(F.XY2LatLon(row[4],row[5],99)[1],8) + '°'.decode('utf-8')
            n = 'N' + '%0.8f' % round(F.XY2LatLon(row[4],row[5],99)[0], 8) + '°'.decode('utf-8')
            x = '%.4f'%round(row[5],4)
            y = '%.4f'%round(row[4],4)
            h = '%.2f'%round(row[6],2)
            if row[7] == 27:
                bz = '电子桩'
            elif row[7] == 2:
                bz = '实体桩'
            elif row[7] == 3:
                bz = '实体桩(移位桩)'
            elif row[7] == 4:
                bz = '告示牌'
            px = int(lc.split('+')[0].replace('K',''))*1000 + int(lc.split('+')[1])
            cn = [name,lc,xzq,e,n,x,y,h,bz,px,pxzy]
            # cnSorted = sorted(cn,key= lambda s:s[9])
            cnList.append(cn)
    cnListSorted = sorted(cnList, key=lambda s:(s[9],s[10]))

    df = pd.DataFrame(cnListSorted,columns=['桩名（编号）','里程','所在位置（地名）','经度','纬度','X','Y','高程（1985国家高程基准）','备注','px','px2'])
    df.to_excel(path.decode('utf-8') + '\\' + ('{}桩牌表.xlsx'.format(riverName.encode('utf-8'))).decode('utf-8') ,index=False)
    xlsx = path + '\\' + '{}桩牌表.xlsx'.format(riverName.encode('utf-8'))
    wb = load_workbook(xlsx.decode('utf-8'))
    ws = wb.worksheets[0]
    ws.insert_rows(0,2)
    fontstyle = Font(size=20, bold=True)
    ws.cell(1, 1, '四川省甘孜藏族自治州巴塘县{}管理线桩信息表'.format(riverName.encode('utf-8')))
    ws['A1'].font = fontstyle
    ws.merge_cells('A1:I1')
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 28
    ws.merge_cells('A2:A3')
    ws.merge_cells('B2:B3')
    ws.merge_cells('C2:C3')
    ws.merge_cells('D2:G2')
    ws.merge_cells('H2:H3')
    ws.merge_cells('I2:I3')
    title_fontstyle = Font(size=12, bold=True)
    ws['A2'] = '桩名（编号）'
    ws['A2'].font = title_fontstyle
    ws['B2'] = '里程'
    ws['B2'].font = title_fontstyle
    ws['C2'] = '所在位置（地名）'
    ws['C2'].font = title_fontstyle
    ws['D2'] = '坐标（2000国家大地坐标系）'
    ws['D2'].font = title_fontstyle
    ws['H2'] = '高程（1985国家高程基准）'
    ws['H2'].font = title_fontstyle
    ws['I2'] = '备注'
    ws['I2'].font = title_fontstyle
    ws.row_dimensions[1].height = 26
    border_set = Border(left=Side(style='thin', color=colors.BLACK),
                        right=Side(style='thin', color=colors.BLACK),
                        top=Side(style='thin', color=colors.BLACK),
                        bottom=Side(style='thin', color=colors.BLACK))
    for row in ws.rows:
        for cell in row:
            cell.border = border_set
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal='center', vertical='center')
    for cell in ws['H']:
        cell.number_format = '0.00'
    wb.save(xlsx.decode('utf-8'))