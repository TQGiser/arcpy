#coding=utf-8
import arcpy
import pandas as pd
import math
from openpyxl import load_workbook
wb = load_workbook(r'D:\2021年项目\0901州河表格-改\湖泊\1.xlsx'.decode('utf-8'))
ws = wb.worksheets[0]
rowNum = []
for row in ws.rows:
    if row[0].value is not None:
        if '告示牌信息表'.decode('utf-8') in row[0].value:
            a =  row[0].row
            rowNum.append(a)
rowNum.sort(reverse=True)
for num in rowNum:
    ws.delete_rows(num,3)
for row in ws.rows:
    if row[0].value is not None:
        if '管理线桩信息表'.decode('utf-8') in row[0].value:
            b = row[0].row
            ws.merge_cells('A{}:H{}'.format(b, b))
wb.save(r'D:\2021年项目\0901州河表格-改\湖泊\all.xlsx'.decode('utf-8'))