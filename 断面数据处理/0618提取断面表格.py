#coding=utf-8
import xlwt
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border,Side,colors,Alignment,Font
df = pd.read_excel(r'D:\2021年项目\0618热曲断面表格处理\工作簿1.xlsx'.decode('utf-8'),header= 0)
df['dmh'] = 0
a = 1
for index,value in df.iterrows():
    if df.loc[index,'点号'.decode('utf-8')] == 1:
        # value['dmh'] = a + 1
        a += 1
    df.loc[index,'dmh'] = a
df.dropna(axis=0,how='any',inplace=True)
df2 = df.groupby('dmh')
for name,group in df2:
    df3 = group
    xlsx = 'RQ-' + '%03.0f'%float(name-1)
    df3.drop(['dmh'], axis=1, inplace=True)
    df3.to_excel(r'D:\2021年项目\0618热曲断面表格处理\新建文件夹\{}.xlsx'.format(xlsx).decode('utf-8'),index=False)
    df4 = pd.read_excel(r'D:\2021年项目\0618热曲断面表格处理\新建文件夹\{}.xlsx'.format(xlsx).decode('utf-8'),header = 1)
    df4.to_excel(r'D:\2021年项目\0618热曲断面表格处理\新建文件夹\{}.xlsx'.format(xlsx).decode('utf-8'),index=False)
    wb = load_workbook((r'D:\2021年项目\0618热曲断面表格处理\新建文件夹\{}.xlsx'.format(xlsx).decode('utf-8')))
    ws = wb.worksheets[0]
    ws.insert_rows(0, 1)
    fontstyle = Font(size= 12,bold=False)
    ws.cell(1,2, '{}'.format(xlsx))
    border_set = Border(left=Side(style='thin', color=colors.BLACK),
                        right=Side(style='thin', color=colors.BLACK),
                        top=Side(style='thin', color=colors.BLACK),
                        bottom=Side(style='thin', color=colors.BLACK))
    for row in ws.rows:
        for cell in row:
            cell.border = border_set

            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = fontstyle
    for cell in ws['E']:
        cell.number_format = '0.00'
    for cell in ws['B']:
        cell.number_format = '0.00'
    for cell in ws['F']:
        cell.number_format = '0.00'
    for cell in ws['C']:
        cell.number_format = '0.00'
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    wb.save(r'D:\2021年项目\0618热曲断面表格处理\新建文件夹\{}.xlsx'.format(xlsx).decode('utf-8'))
