# coding=utf-8
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border,Side,colors,Alignment
path = r'D:\2021年项目\0429断面采集\处理成果\RQ-001.xlsx'.decode('utf-8')
A = '执曲横断面测量成果表'.decode('utf-8')
df = pd.read_excel(path,header= None,names = [A,'','','','',''])
outpath = r'D:\2021年项目\0429断面采集\处理成果\RQ-001.xlsx'.decode('utf-8')
df.to_excel(outpath,index=False)
exceldress = r'D:\2021年项目\0429断面采集\处理成果\RQ-001.xlsx'.decode('utf-8')
wb = load_workbook(exceldress)
ws = wb.worksheets[0]
border_set = Border(left=Side(style='thin',color=colors.BLACK),
                    right=Side(style='thin',color=colors.BLACK),
                    top=Side(style='thin',color=colors.BLACK),
                    bottom=Side(style='thin',color=colors.BLACK))
for row in ws.rows:
    for cell in row:
        cell.border = border_set
        cell.alignment = Alignment(horizontal='center',vertical='center')
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 12
ws.merge_cells('A1:F1')
wb.save(r'D:\2021年项目\0429断面采集\处理成果\RQ-001.xlsx'.decode('utf-8'))
