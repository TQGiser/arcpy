# coding=utf-8
import pandas as pd
import matplotlib
import math
import os
from openpyxl import load_workbook
from openpyxl.styles import Border,Side,colors,Alignment
A = '执曲横断面测量成果表'.decode('utf-8')
path = r'D:\2021年项目\0705拉合库断面读取\表格\bb.xlsx'.decode('utf-8')
outpath = r'D:\2021年项目\0705拉合库断面读取\表格\OK'.decode('utf-8')
df = pd.read_excel(path)
df['pj'] = 0
df['ELEV'] = df['ELEV'].round(2)
df['Y'] = df['Y'].round(4)
df['X'] = df['X'].round(4)
df['pj'] = df['pj'].round(2)
df2 = df.groupby('dmh')
for name,group in df2:
    x0 = group[group['Qd'] == 'qd'.decode('utf-8')]['X']
    y0 = group[group['Qd'] == 'qd'.decode('utf-8')]['Y']
    for index,value in group.iterrows():
        value['pj'] = round(math.sqrt((value['X'] - x0) ** 2 + (value['Y'] - y0) ** 2),2)
        df.loc[index, 'pj'] = value['pj']
df3 = df.groupby('dmh')
for name,group in df3:
    df4 = group
    df4.sort_values(by = 'pj',ascending= True,inplace = True)
    df4['dh'] = df4['pj'].rank()
    # df4.drop(['OBJECTID', 'dm', 'QD','lc'], axis=1, inplace=True)
    dms = pd.DataFrame({'pj':[name]})
    df5 = pd.concat([dms,df4],keys = 'dh',ignore_index= True)
    df5 = df5[['dh','pj','ELEV','Sbd','X','Y']]
    df5.columns = ['点号'.decode('utf-8'),'另点距'.decode('utf-8'),'高程'.decode('utf-8'),'备注'.decode('utf-8'),'X','Y']
    filename = outpath + '\\' + name + '.xlsx'
    df5.to_excel(filename, index=False)
    print '{}'.format(name) + '断面处理完成'
os.chdir(outpath)
lst = []
for i,v,m in os.walk(outpath):
    lst=m
def get_dt(lst):
    for i in lst:
        df = pd.read_excel(i)
        yield df
print '正在合并所有断面数据'
df = pd.concat(get_dt(lst))
allexcel = outpath + '\\' + 'all.xlsx'
df.to_excel(allexcel,index = False)
lst2 = []
for i,v,m in os.walk(outpath):
    lst2=m
print '正在设备表格边框及表格标题'
for dmfile in lst2:
    df = pd.read_excel(dmfile, header=None, names=[A, '', '', '', '', ''])
    addTitleFile = outpath + '\\' +dmfile
    df.to_excel(addTitleFile, index=False)
    wb = load_workbook(addTitleFile)
    ws = wb.worksheets[0]
    border_set = Border(left=Side(style='thin', color=colors.BLACK),
                        right=Side(style='thin', color=colors.BLACK),
                        top=Side(style='thin', color=colors.BLACK),
                        bottom=Side(style='thin', color=colors.BLACK))
    for row in ws.rows:
        for cell in row:
            cell.border = border_set
            cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.merge_cells('A1:F1')
    setStlyeFile = outpath + '\\' + dmfile
    wb.save(setStlyeFile)
df = pd.read_excel(allexcel,header=1)
df.drop(['X','Y','备注'.decode('utf-8')], axis=1, inplace=True)
a = df[df.T.isnull().any()]
print '正在处理cass格式'
for index, value in a.iterrows():
    df.loc[index, '高程'.decode('utf-8')] = value['另点距'.decode('utf-8')]
    df.loc[index,'另点距'.decode('utf-8')] = 'BEGIN'
df.drop(['点号'.decode('utf-8')],axis=1,inplace= True)
cassfile = outpath + '\\' + 'cass.xls'
df.to_excel(cassfile,index = False,header=False)